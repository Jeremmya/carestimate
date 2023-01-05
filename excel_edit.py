from openpyxl import load_workbook
import main
#import os
#print('Get current working directory : ', os.getcwd())

def excel_editor(price1,price2,price3, model, mile_or, year_or, mile_1, year_1, mile_2, year_2, mile_3, year_3):
	#https://dl.uploadgram.me/63b5780d76ca4h
	#load excel file
	#workbook = load_workbook(filename="/Users/mike13/Downloads/starter.xlsx")
	workbook = load_workbook(filename="https://dl.uploadgram.me/63b5780d76ca4h")

	kurs = main.kurs_valyut()
	 
	#open workbook
	sheet = workbook.active

	year_1s = year_1.replace('Год выпуска: ', '')
	year_2s = year_2.replace('Год выпуска: ', '')
	year_3s = year_3.replace('Год выпуска: ', '')
	print(price1)
	price1s = price1.replace('у.е.', '')
	print(price1s)
	price1se = price1s.replace(' ', '')
	price2s = price2.replace('у.е.', '')
	price2se = price2s.replace(' ', '')
	price3s = price3.replace('у.е.', '')
	price3se = price3s.replace(' ', '')
	print(price1se)

	 
	#modify the desired cell
	sheet["H16"] = int(price1se) #цена конкурентного авто 1
	sheet["I16"] = int(price2se) #цена конкурентного авто 2
	sheet["J16"] = int(price3se) #цена конкурентного авто 3
	sheet["B2"] = model # наименование авто
	sheet["A2"] = "Full Name" # гос номер
	sheet["M17"] = int(mile_or/1000) # пробег оригинал авто
	sheet["M18"] = int(year_or) # год выпуска оригинал авто
	sheet["N17"] = int(mile_1/1000)# пробег 1 авто
	sheet["N18"] = int(year_1s) # год выпуска 1 авто
	sheet["O17"] = int(mile_2/1000) # пробег 2 авто
	sheet["O18"] = int(year_2s) # год выпуска 2 авто
	sheet["P17"] = int(mile_3/1000) # пробег 3 авто
	sheet["P18"] = int(year_3s) # год выпуска 3 
	sheet["L12"]= int(kurs)

	if year_1s.isnumeric():
		print("is numeric")
	 
	#save the file
	workbook.save(filename="https://dl.uploadgram.me/63b5780d76ca4h")



if __name__ == '__main__':
	excel_editor()